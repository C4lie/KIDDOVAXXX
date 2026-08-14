import threading
import os
from django.conf import settings
from openpyxl import Workbook

excel_lock = threading.Lock()

def export_db_to_excel():
    """
    Exports vaccinations, hospitals, receptionists, patients, and admins database records
    to a multi-sheet Excel file 'hospital_credentials.xlsx' in the project root directory.
    """
    with excel_lock:
        # Import models inside the function to prevent circular imports during startup
        from adminapp.models import Admintbl
        from hospitalapp.models import Hospitaltbl, Vaccinetbl, Receptionisttbl
        from patientapp.models import Patienttbl

        wb = Workbook()

        # 1. Vaccinations & Prices Sheet
        ws_vaccines = wb.active
        ws_vaccines.title = "Vaccinations"
        ws_vaccines.append([
            "Vaccine ID", 
            "Vaccine Name", 
            "Price", 
            "Stock Quantity", 
            "Minimum Quantity", 
            "Hospital Title", 
            "Hospital ID"
        ])
        for v in Vaccinetbl.objects.all().select_related('hospitalId'):
            h_title = v.hospitalId.title if v.hospitalId else "N/A"
            h_id = v.hospitalId.id if v.hospitalId else "N/A"
            ws_vaccines.append([
                v.id, 
                v.vaccineName, 
                v.price, 
                v.stock_quantity, 
                v.minimum_quantity, 
                h_title, 
                h_id
            ])

        # 2. Hospitals Sheet
        ws_hospitals = wb.create_sheet(title="Hospitals")
        ws_hospitals.append([
            "Hospital ID", 
            "Hospital Title", 
            "Contact (Username)", 
            "Password", 
            "Doctor Name", 
            "Address", 
            "City", 
            "Area"
        ])
        for h in Hospitaltbl.objects.all().select_related('cityId', 'areaId'):
            c_name = h.cityId.cityName if h.cityId else "N/A"
            a_name = h.areaId.areaName if h.areaId else "N/A"
            ws_hospitals.append([
                h.id, 
                h.title, 
                h.contactNo, 
                h.password, 
                h.dcrname, 
                h.address, 
                c_name, 
                a_name
            ])

        # 3. Receptionists Sheet
        ws_receptionists = wb.create_sheet(title="Receptionists")
        ws_receptionists.append([
            "Receptionist ID", 
            "Name", 
            "Contact (Username)", 
            "Password", 
            "Hospital Title", 
            "Hospital ID", 
            "Gender", 
            "Address", 
            "Date of Joining"
        ])
        for r in Receptionisttbl.objects.all().select_related('hospitalid'):
            h_title = r.hospitalid.title if r.hospitalid else "N/A"
            h_id = r.hospitalid.id if r.hospitalid else "N/A"
            doj_str = r.doj.strftime("%Y-%m-%d") if r.doj else "N/A"
            ws_receptionists.append([
                r.id, 
                r.name, 
                r.contactNo, 
                r.password, 
                h_title, 
                h_id, 
                r.gender, 
                r.address, 
                doj_str
            ])

        # 4. Patients Sheet
        ws_patients = wb.create_sheet(title="Patients")
        ws_patients.append([
            "Patient ID", 
            "Name", 
            "Contact (Username)", 
            "Password", 
            "Address", 
            "City", 
            "Area"
        ])
        for p in Patienttbl.objects.all().select_related('cityId', 'areaId'):
            c_name = p.cityId.cityName if p.cityId else "N/A"
            a_name = p.areaId.areaName if p.areaId else "N/A"
            ws_patients.append([
                p.id, 
                p.name, 
                p.contactNo, 
                p.password, 
                p.address, 
                c_name, 
                a_name
            ])

        # 5. Admins Sheet
        ws_admins = wb.create_sheet(title="Admins")
        ws_admins.append([
            "Admin ID", 
            "Username", 
            "Password"
        ])
        for a in Admintbl.objects.all():
            ws_admins.append([
                a.id, 
                a.username, 
                a.password
            ])

        # Path to save
        file_path = os.path.join(settings.BASE_DIR, 'hospital_credentials.xlsx')
        
        # Save to temporary file first, then rename atomically to prevent corruption/locks
        temp_path = file_path + '.tmp'
        try:
            wb.save(temp_path)
            if os.path.exists(file_path):
                os.remove(file_path)
            os.rename(temp_path, file_path)
        except Exception as e:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise e
